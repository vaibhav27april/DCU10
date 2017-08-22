import openpyxl
global_signal_data=''
global_message_data=''
old_global_message_data=''
x=0
local_i=0
sheet=''
my_list=[
        ['a',1,2,3,4],
        ['b',5,6,7,8]
        ]
file1 = open("MyFile1.can","w")

my_try_list=[[]]
time_list=[]
# def extract_data():
#     global global_signal_data
#     global global_message_data
#     global old_global_message_data
#     global my_try_list
#     check_j=0
#     j = 0
#     global_signal_data = ''
#     global_message_data = ''
#     wb = openpyxl.load_workbook('170510_ADAS_DRV_ARXML_IF rev8 VB.xlsx')
#     type(wb)
#     sheet = wb.get_sheet_by_name("ADAS 통합제어")
#     for i in range(5, 45):
#      global_signal_data = (sheet['J' + str(i)].value)
#      global_message_data = (sheet['H' + str(i)].value)
#      if global_message_data != old_global_message_data:
#             if check_j==1:
#              j = j + 1                             #required to create 2D list for each new message name
#              my_try_list = my_try_list + [[]]    #create 2D list at run time
#             my_try_list[j].append(global_message_data)
#             old_global_message_data = global_message_data
#
#
#      my_try_list[j].append(global_signal_data)
#      check_j=1
#      #j = j + 1
#     print(my_try_list)
#     return
#
#
#
#
#
#
#
# extract_data()


def readxls_file(key,readxlx_local_i,readxlx_local_j):

     global global_signal_data
     global global_message_data
     global sheet
     global_signal_data=''
     global_message_data=''

     wb = openpyxl.load_workbook('170510_ADAS_DRV_ARXML_IF rev8 VB.xlsx')
     type(wb)
     sheet=wb.get_sheet_by_name("ADAS 통합제어")

     for i in range(readxlx_local_i,readxlx_local_j):
      global_signal_data = (sheet['J'+str(i)].value)
      global_message_data = (sheet['H' + str(i)].value)
      dict1[key]()
      print(global_signal_data)

     #return



def a():
    global global_signal_data
    global global_message_data
    global old_global_message_data

    if global_message_data!=old_global_message_data:
     file1.write('message ' + global_message_data +" "+ global_message_data + ';\n')
     old_global_message_data=global_message_data
    return



def b():
    global global_signal_data
    global global_message_data
    global old_global_message_data
    if global_message_data != old_global_message_data:
        file1.write('setData' + global_message_data + '()' + ';\n')

        old_global_message_data = global_message_data
    return



def c():
    global global_signal_data
    global global_message_data
    global old_global_message_data
    if global_message_data != old_global_message_data:
        file1.write('output(' + global_message_data + ')' + ';\n')
        old_global_message_data = global_message_data
    return



def d():

    global global_signal_data
    global global_message_data
    global old_global_message_data
    global x
    if global_message_data != old_global_message_data:
        if x==1:
            file1.write('}\n')
            x=0
        file1.write('void ' + 'setData' + global_message_data + '(void)\n')
        file1.write('{\n')
        x=1
        old_global_message_data = global_message_data
    file1.write(global_message_data + '.' + global_signal_data + '=' + '@' + 'ENV_' + global_signal_data + ';\n')
    return



def e():
    file1.write(global_message_data + '.' + global_signal_data + '=' + '@' + 'ENV_' + global_signal_data +';\n')
    return

def f():
    global global_signal_data
    global global_message_data
    global old_global_message_data
    file1.write('@ENV_'+global_signal_data+'_IVC_R' + '=@Sysvar::XCP::XCPA2L::'+global_signal_data+'_IVC_R;\n')
    return

dict1={'1':a,'2':b,'3':c,'4':d,'5':e,'6':f}


def extract_time():
    global time_list
    global sheet
    old_global_time_data = 0
    New_global_time_data = 0
    wb = openpyxl.load_workbook('170510_ADAS_DRV_ARXML_IF rev8 VB.xlsx')
    type(wb)
    sheet = wb.get_sheet_by_name("ADAS 통합제어")
    for time_i in range(5, 46):#lenth from 5 to 46 isrequired to get list diffrence of blank value
        New_global_time_data = (sheet['N' + str(time_i)].value)
        if old_global_time_data != New_global_time_data:
            time_list.append(time_i)
            old_global_time_data = New_global_time_data
    return



def timer_generation(timer_data):
    global local_i
    global time_list
    wb = openpyxl.load_workbook('170510_ADAS_DRV_ARXML_IF rev8 VB.xlsx')
    type(wb)
    sheet = wb.get_sheet_by_name("ADAS 통합제어")
    xyz=time_list[local_i]
    value_l = sheet['N' + str(xyz)].value
    file1.write('on timer t' + str(timer_data) + '\n')
    file1.write('{\n')
    readxls_file('2', time_list[local_i], time_list[(local_i + 1)])
    # file1.write('output('+'ADAS_CMD_30_10ms'+'_'+')'+';\n')
    readxls_file('3', time_list[local_i], time_list[(local_i + 1)])
    readxls_file('6', time_list[local_i], time_list[(local_i + 1)])
    file1.write('setTimer(t' + str(timer_data) + ',' + str(value_l) + ');\n')
    file1.write('}\n')
    return
#readxls_file('3')
extract_time()
timer_data_local_len=len(time_list)

file1.write('/*@!Encoding:949*/\n')
file1.write('{\n')
file1.write('}\n')

file1.write('variables\n')
file1.write('{\n')
#file1.write('message '+'ADAS_CMD_30_10ms '+'ADAS_CMD_30_10ms_'+';\n')
readxls_file('1',5,45)
file1.write('mstimer  ' + 't1'+';\n')
file1.write('}\n')

file1.write('on start\n')
file1.write('{\n')
for settimer_l in range(1,timer_data_local_len):
 file1.write('setTimer(t'+str(settimer_l)+',0);\n')#need to check how many timer required

file1.write('}\n')


for localincrement_i in range(1,timer_data_local_len):
    timer_generation(localincrement_i)
    local_i=local_i+1

# file1.write('on timer t1\n')
# file1.write('{\n')
# #file1.write('setData'+'ADAS_CMD_30_10ms'+'()'+';\n')
# readxls_file('2',5,45)
# #file1.write('output('+'ADAS_CMD_30_10ms'+'_'+')'+';\n')
# readxls_file('3',5,45)
# readxls_file('6',5,45)
# file1.write('setTimer(t1,10);\n')
# file1.write('}\n')

#file1.write('void '+'setData'+'ADAS_CMD_30_10ms'+'(void)')
readxls_file('4',5,45)

#file1.write('ADAS_CMD_30_10ms_'+'.'+'ADAS_CMD_30_Crc01Val'+'='+'@'+'ENV_'+'ADAS_CMD_30_Crc01Val\n')
#readxls_file('5')
file1.write('}\n')

